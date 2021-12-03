import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
import os

wb = Workbook()
sheet1 = wb.create_sheet("BigSheet1", 0)
print(wb.sheetnames)
print(sheet1['A1'].value)
sheet1['A1'] = 'Hello'
sheet1['A2'] = 42

os.chdir('some/path/here')
wb.save('bad_excel_example.xlsx')

wb2 = load_workbook('bad_excel_example.xlsx')
print(wb2.sheetnames)
wb2Sheet = wb2['BigSheet1']
print(wb2Sheet['A1'].value)
