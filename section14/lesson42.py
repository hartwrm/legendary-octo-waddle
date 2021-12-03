import openpyxl
import os
os.chdir('/your/path/here')
from openpyxl import load_workbook

workbook = load_workbook('example_excel.xlsx')
print(workbook.sheetnames)
sheet = workbook['Sheet1']
print(sheet['A1'].value)

for i in range (1, 8):
    print(i, sheet.cell(row=i, column=2).value)
